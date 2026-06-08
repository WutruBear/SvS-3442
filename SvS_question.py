import streamlit as st
import pandas as pd
import re
import io
from datetime import datetime

import networkx as nx

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

st.set_page_config(page_title="SvS #3442 tool", page_icon="⚔️", layout="wide")

# ═══════════════════════════════════════════════════════════════════════════════
# CSS
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&family=Barlow:wght@400;500;600;700&family=Barlow+Condensed:wght@500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

    /* ── TOKENS ──────────────────────────────────────────────── */
    :root {
        --ink:        #05080f;
        --surface-0: #080c14;
        --surface-1: #0c1119;
        --surface-2: #101620;
        --surface-3: #141c28;
        --line:      rgba(255,255,255,0.06);
        --line-2:    rgba(255,255,255,0.10);
        --text:      #b8c8dc;
        --text-2:    #6a7e96;
        --text-3:    #364452;
        --gold:      #c8a84b;
        --gold-dim:  rgba(200,168,75,0.12);
        --gold-line: rgba(200,168,75,0.28);
        --blue:      #3d7eff;
        --blue-glow: rgba(61,126,255,0.15);
        --blue-line: rgba(61,126,255,0.30);
        --green:     #2db882;
        --green-dim: rgba(45,184,130,0.10);
        --green-line:rgba(45,184,130,0.28);
        --amber:     #d4893a;
        --amber-dim: rgba(212,137,58,0.10);
        --amber-line:rgba(212,137,58,0.28);
        --red:       #c94040;
        --red-dim:   rgba(201,64,64,0.10);
        --red-line:  rgba(201,64,64,0.28);
        --r:         8px;
        --r-sm:      5px;
    }

    /* ── BASE ────────────────────────────────────────────────── */
    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
        -webkit-font-smoothing: antialiased;
    }
    .stApp {
        background: var(--ink);
        background-image:
            radial-gradient(ellipse 80% 40% at 50% -10%, rgba(61,126,255,0.07) 0%, transparent 70%),
            radial-gradient(ellipse 60% 30% at 80% 100%, rgba(200,168,75,0.04) 0%, transparent 60%);
        color: var(--text);
    }
    #MainMenu, footer, header { visibility: hidden; }
    h1, h2, h3, h4 {
        font-family: 'Barlow', sans-serif;
        font-weight: 700; color: #d4dfe8; letter-spacing: -0.02em;
    }
    label {
        color: var(--text-2) !important;
        font-size: 0.72rem !important;
        font-weight: 600 !important;
        text-transform: uppercase !important;
        letter-spacing: 0.09em !important;
        font-family: 'Barlow', sans-serif !important;
    }
    section[data-testid="stSidebar"],
    button[data-testid="collapsedControl"] { display: none !important; }
    .block-container {
        padding-top: 0 !important;
        padding-left: 2rem !important;
        padding-right: 2rem !important;
        max-width: 1240px !important;
    }
    @media (max-width: 640px) {
        .block-container {
            padding-left: 1rem !important;
            padding-right: 1rem !important;
        }
    }

    /* ── SECTION LABEL ───────────────────────────────────────── */
    .section-label {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 0.65rem; font-weight: 600; text-transform: uppercase;
        letter-spacing: 0.18em; color: var(--text-3); margin-bottom: 0.6rem;
    }

    /* ── TOP NAV ─────────────────────────────────────────────── */
    .topnav {
        position: relative; overflow: hidden;
        display: flex; align-items: stretch;
        background: var(--surface-1);
        border-bottom: 1px solid var(--line-2);
        margin: 0 -2rem 2rem -2rem;
        padding: 0;
    }
    .topnav::after {
        content: '';
        position: absolute; bottom: 0; left: 0; right: 0; height: 1px;
        background: linear-gradient(90deg,
            transparent 0%,
            var(--gold) 20%,
            var(--gold) 80%,
            transparent 100%);
        opacity: 0.25;
    }
    .topnav-inner {
        display: flex; align-items: stretch;
        justify-content: space-between;
        width: 100%; max-width: 1240px;
        margin: 0 auto;
        padding: 0 2rem;
    }
    .topnav-brand {
        display: flex; align-items: center; gap: 0.9rem;
        padding: 1rem 0; flex-shrink: 0;
        border-right: 1px solid var(--line);
        padding-right: 1.6rem; margin-right: 0;
    }
    .topnav-icon {
        width: 34px; height: 34px; flex-shrink: 0;
        display: flex; align-items: center; justify-content: center;
        background: var(--gold-dim);
        border: 1px solid var(--gold-line);
        border-radius: var(--r-sm);
    }
    .topnav-icon svg { width: 18px; height: 18px; }
    .topnav-text { display: flex; flex-direction: column; gap: 1px; }
    .topnav-name {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 1.2rem; font-weight: 700;
        line-height: 1; letter-spacing: 0.08em; text-transform: uppercase;
        color: #dce6f0;
    }
    .topnav-name em { font-style: normal; color: var(--gold); }
    .topnav-tagline {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 0.58rem; font-weight: 600;
        letter-spacing: 0.22em; text-transform: uppercase;
        color: var(--text-3);
    }
    .topnav-right {
        display: flex; align-items: center; gap: 1rem;
        padding: 1rem 0; padding-left: 1.6rem;
    }
    .topnav-badge {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 0.7rem; font-weight: 600;
        letter-spacing: 0.12em; text-transform: uppercase;
        color: var(--green); background: var(--green-dim);
        border: 1px solid var(--green-line);
        border-radius: var(--r-sm);
        padding: 0.25rem 0.75rem; white-space: nowrap;
    }
    .topnav-badge.dim {
        color: var(--text-3); background: transparent;
        border-color: var(--line);
    }
    .topnav-sep { width: 1px; height: 18px; background: var(--line); }
    .topnav-meta {
        font-family: 'JetBrains Mono', monospace;
        font-size: 0.6rem; color: var(--text-3); letter-spacing: 0.04em;
    }
    @media (max-width: 560px) {
        .topnav { margin-left: -1rem; margin-right: -1rem; }
        .topnav-inner { padding: 0 1rem; }
        .topnav-tagline, .topnav-meta, .topnav-sep { display: none; }
        .topnav-name { font-size: 1rem; }
        .topnav-brand { padding-right: 1rem; }
        .topnav-right { padding-left: 1rem; gap: 0.6rem; }
    }

    /* ── STEPPER ─────────────────────────────────────────────── */
    .stepper {
        display: flex; align-items: center;
        background: transparent;
        border: 1px solid var(--line);
        border-radius: var(--r);
        padding: 0.7rem 1.1rem;
        margin-bottom: 1.75rem;
        overflow-x: auto; gap: 0;
        scrollbar-width: none;
    }
    .stepper::-webkit-scrollbar { display: none; }
    .step { display: flex; align-items: center; gap: 0.5rem; flex-shrink: 0; }
    .step-circle {
        width: 22px; height: 22px;
        border-radius: 3px;
        display: flex; align-items: center; justify-content: center;
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 0.7rem; font-weight: 700; flex-shrink: 0;
        transition: all 0.2s;
    }
    .step-circle.done   { background: var(--green-dim); color: var(--green); border: 1px solid var(--green-line); }
    .step-circle.active { background: var(--blue-glow); color: var(--blue); border: 1px solid var(--blue-line); box-shadow: 0 0 10px rgba(61,126,255,0.2); }
    .step-circle.idle   { background: var(--surface-2); color: var(--text-3); border: 1px solid var(--line); }
    .step-label {
        font-family: 'Barlow', sans-serif;
        font-size: 0.75rem; font-weight: 600; white-space: nowrap;
    }
    .step-label.done   { color: var(--green); }
    .step-label.active { color: #7aaef8; }
    .step-label.idle   { color: var(--text-3); }
    .step-arrow { color: var(--text-3); font-size: 0.65rem; margin: 0 0.65rem; flex-shrink: 0; }
    @media (max-width: 600px) {
        .step-label { display: none; }
        .step-circle { width: 26px; height: 26px; font-size: 0.72rem; }
        .step-arrow { margin: 0 0.4rem; }
    }

    /* ── STAT CARDS ──────────────────────────────────────────── */
    .stats-row {
        display: grid; grid-template-columns: repeat(4, 1fr);
        gap: 0.75rem; margin-bottom: 1.75rem;
    }
    @media (max-width: 640px) { .stats-row { grid-template-columns: repeat(2, 1fr); } }
    @media (max-width: 340px) { .stats-row { grid-template-columns: 1fr; } }
    .stat-card {
        background: var(--surface-1);
        border: 1px solid var(--line);
        border-radius: var(--r);
        padding: 1.1rem 1.2rem;
        position: relative; overflow: hidden;
        transition: border-color 0.2s, transform 0.15s;
    }
    .stat-card::before {
        content: ''; position: absolute;
        top: 0; left: 0; right: 0; height: 2px;
        background: linear-gradient(90deg, var(--blue), transparent);
        opacity: 0.4;
    }
    .stat-card:hover { border-color: var(--line-2); transform: translateY(-1px); }
    .stat-card .stat-value {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 2rem; font-weight: 700; color: #d4dfe8;
        line-height: 1; margin-bottom: 0.35rem; letter-spacing: -0.02em;
    }
    .stat-card .stat-label {
        font-family: 'Barlow', sans-serif;
        font-size: 0.65rem; font-weight: 600; text-transform: uppercase;
        letter-spacing: 0.1em; color: var(--text-3);
    }
    .stat-card.warn {
        border-color: var(--amber-line);
        background: var(--amber-dim);
    }
    .stat-card.warn::before { background: linear-gradient(90deg, var(--amber), transparent); opacity: 0.5; }
    .stat-card.warn .stat-value { color: var(--amber); }

    /* ── BANNERS ─────────────────────────────────────────────── */
    .alert-banner, .error-banner, .info-banner, .success-banner {
        border-radius: var(--r-sm);
        padding: 0.85rem 1rem 0.85rem 1.1rem;
        margin: 0.7rem 0;
        font-family: 'Inter', sans-serif;
        font-size: 0.82rem;
        line-height: 1.55;
        position: relative;
    }
    .alert-banner {
        background: var(--amber-dim);
        border: 1px solid var(--amber-line);
        border-left: 3px solid var(--amber);
        color: #c89848;
    }
    .error-banner {
        background: var(--red-dim);
        border: 1px solid var(--red-line);
        border-left: 3px solid var(--red);
        color: #c07070;
    }
    .info-banner {
        background: var(--blue-glow);
        border: 1px solid var(--blue-line);
        border-left: 3px solid var(--blue);
        color: #7aaef8;
    }
    .success-banner {
        background: var(--green-dim);
        border: 1px solid var(--green-line);
        border-left: 3px solid var(--green);
        color: #4aaf88;
    }
    .field-warn {
        font-size: 0.7rem; color: var(--amber);
        margin-top: -0.25rem; margin-bottom: 0.5rem;
        font-family: 'Barlow', sans-serif; font-weight: 500;
    }

    /* ── FORM INPUTS ─────────────────────────────────────────── */
    textarea {
        font-family: 'JetBrains Mono', monospace !important;
        font-size: 0.77rem !important; line-height: 1.75 !important;
        background: var(--surface-1) !important; color: #8aa8c4 !important;
        border: 1px solid var(--line) !important;
        border-radius: var(--r-sm) !important;
    }
    textarea:focus {
        border-color: var(--blue-line) !important;
        box-shadow: 0 0 0 3px var(--blue-glow) !important;
    }
    input[type="text"], .stTextInput input, .stNumberInput input {
        background: var(--surface-2) !important;
        border: 1px solid var(--line) !important;
        border-radius: var(--r-sm) !important;
        color: var(--text) !important;
        font-family: 'Inter', sans-serif !important;
        font-size: 0.85rem !important;
        min-height: 42px !important;
    }
    input:focus {
        border-color: var(--blue-line) !important;
        box-shadow: 0 0 0 3px var(--blue-glow) !important;
    }
    .stSelectbox > div > div {
        background: var(--surface-2) !important;
        border: 1px solid var(--line) !important;
        border-radius: var(--r-sm) !important;
        color: var(--text) !important;
        min-height: 42px !important;
    }
    .stMultiSelect > div > div {
        background: var(--surface-2) !important;
        border: 1px solid var(--line) !important;
        border-radius: var(--r-sm) !important;
    }

    /* ── BUTTONS ─────────────────────────────────────────────── */
    .stButton > button, .stDownloadButton > button, .stFormSubmitButton > button {
        background: var(--surface-2);
        color: var(--text-2);
        border: 1px solid var(--line-2);
        border-radius: var(--r-sm);
        font-family: 'Barlow', sans-serif;
        font-weight: 600;
        font-size: 0.82rem;
        letter-spacing: 0.03em;
        padding: 0.55rem 1.2rem;
        min-height: 42px;
        transition: all 0.15s;
    }
    .stButton > button:hover, .stDownloadButton > button:hover {
        background: var(--surface-3);
        border-color: rgba(255,255,255,0.14);
        color: var(--text);
        box-shadow: 0 2px 12px rgba(0,0,0,0.35);
    }
    .stButton > button:active { transform: translateY(1px); }
    .stButton > button[kind="primary"] {
        background: rgba(61,126,255,0.12) !important;
        color: #7aaef8 !important;
        border-color: var(--blue-line) !important;
        box-shadow: inset 0 1px 0 rgba(255,255,255,0.04) !important;
    }
    .stButton > button[kind="primary"]:hover {
        background: rgba(61,126,255,0.20) !important;
        border-color: rgba(61,126,255,0.5) !important;
        color: #a8c8ff !important;
        box-shadow: 0 0 20px rgba(61,126,255,0.15) !important;
    }
    @media (max-width: 480px) {
        .stButton > button, .stDownloadButton > button { font-size: 0.78rem; }
    }

    /* ── EXPANDER ────────────────────────────────────────────── */
    .streamlit-expanderHeader {
        background: var(--surface-1) !important;
        border: 1px solid var(--line) !important;
        border-radius: var(--r-sm) !important;
        color: var(--text-2) !important;
        font-family: 'Barlow', sans-serif !important;
        font-size: 0.83rem !important;
        font-weight: 600 !important;
        padding: 0.7rem 0.95rem !important;
        letter-spacing: 0.01em !important;
    }
    .streamlit-expanderHeader:hover {
        border-color: var(--line-2) !important;
        color: var(--text) !important;
    }
    .streamlit-expanderContent {
        background: var(--surface-1) !important;
        border: 1px solid var(--line) !important;
        border-top: none !important;
        border-radius: 0 0 var(--r-sm) var(--r-sm) !important;
    }

    /* ── DATAFRAME ───────────────────────────────────────────── */
    .stDataFrame {
        border: 1px solid var(--line) !important;
        border-radius: var(--r) !important;
        overflow: hidden;
    }
    .stDataFrame > div { overflow-x: auto !important; }

    /* ── TABS ────────────────────────────────────────────────── */
    .stTabs [data-baseweb="tab-list"] {
        background: transparent !important;
        gap: 0.2rem;
        border-bottom: 1px solid var(--line) !important;
    }
    .stTabs [data-baseweb="tab"] {
        background: transparent !important;
        color: var(--text-3) !important;
        font-family: 'Barlow', sans-serif !important;
        font-size: 0.8rem !important;
        font-weight: 600 !important;
        border-radius: var(--r-sm) var(--r-sm) 0 0 !important;
        border: 1px solid transparent !important;
        border-bottom: none !important;
        padding: 0.5rem 1rem !important;
        transition: all 0.15s !important;
        letter-spacing: 0.02em !important;
    }
    .stTabs [data-baseweb="tab"]:hover { color: var(--text-2) !important; }
    .stTabs [aria-selected="true"] {
        background: var(--surface-2) !important;
        color: var(--text) !important;
        border-color: var(--line) !important;
        border-bottom-color: var(--surface-2) !important;
    }

    /* ── RADIO ───────────────────────────────────────────────── */
    .stRadio [data-testid="stWidgetLabel"] { color: var(--text-2) !important; }
    .stRadio label {
        text-transform: none !important;
        letter-spacing: 0 !important;
        font-size: 0.85rem !important;
        font-weight: 500 !important;
        color: var(--text) !important;
    }

    /* ── METRICS ─────────────────────────────────────────────── */
    [data-testid="metric-container"] {
        background: var(--surface-1);
        border: 1px solid var(--line);
        border-radius: var(--r);
        padding: 0.85rem 1.1rem;
    }
    [data-testid="metric-container"] label {
        font-size: 0.68rem !important;
        color: var(--text-3) !important;
        font-family: 'Barlow', sans-serif !important;
        text-transform: uppercase !important;
        letter-spacing: 0.08em !important;
    }
    [data-testid="metric-container"] [data-testid="stMetricValue"] {
        color: #d4dfe8 !important;
        font-family: 'Barlow Condensed', sans-serif !important;
        font-weight: 700 !important;
        font-size: 1.6rem !important;
    }

    /* ── FC-RAW BLOCK ────────────────────────────────────────── */
    .fc-raw {
        font-family: 'JetBrains Mono', monospace;
        font-size: 0.74rem; color: var(--text-2);
        margin-bottom: 0.9rem; padding: 0.6rem 0.85rem;
        background: var(--surface-0);
        border-radius: var(--r-sm);
        border: 1px solid var(--line);
    }
    .fc-raw span { color: #5a9ad8; }

    /* ── DIVIDER ─────────────────────────────────────────────── */
    hr {
        border: none;
        border-top: 1px solid var(--line);
        margin: 1.75rem 0;
    }

    /* ── GUIDE CARDS ─────────────────────────────────────────── */
    .guide-grid {
        display: grid; grid-template-columns: repeat(3, 1fr);
        gap: 1rem; margin-bottom: 1.5rem;
    }
    @media (max-width: 700px) { .guide-grid { grid-template-columns: 1fr; } }
    .guide-card {
        background: var(--surface-1);
        border: 1px solid var(--line);
        border-radius: var(--r);
        padding: 1.1rem 1.2rem;
        transition: border-color 0.2s;
        position: relative; overflow: hidden;
    }
    .guide-card::after {
        content: ''; position: absolute;
        top: 0; left: 0; bottom: 0; width: 2px;
        background: var(--gold);
        opacity: 0.3;
    }
    .guide-card:hover { border-color: var(--line-2); }
    .guide-card-icon {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 0.65rem; font-weight: 700;
        letter-spacing: 0.18em; text-transform: uppercase;
        color: var(--gold); margin-bottom: 0.5rem;
    }
    .guide-card-title {
        font-family: 'Barlow', sans-serif;
        font-size: 0.88rem; font-weight: 700;
        color: #d4dfe8; margin-bottom: 0.35rem;
    }
    .guide-card-body {
        font-size: 0.78rem; color: var(--text-2);
        line-height: 1.6;
    }
    .guide-card-body code {
        font-family: 'JetBrains Mono', monospace; font-size: 0.7rem;
        color: #5a9ad8; background: var(--surface-0);
        padding: 0.1rem 0.35rem; border-radius: 3px;
    }

    /* ── FORMAT TABLE ────────────────────────────────────────── */
    .fmt-table { width: 100%; border-collapse: collapse; font-size: 0.78rem; }
    .fmt-table th {
        text-align: left; padding: 0.5rem 0.8rem;
        color: var(--text-3); font-family: 'Barlow', sans-serif;
        font-size: 0.65rem; text-transform: uppercase; letter-spacing: 0.1em;
        border-bottom: 1px solid var(--line);
    }
    .fmt-table td {
        padding: 0.42rem 0.8rem; color: var(--text-2);
        border-bottom: 1px solid rgba(255,255,255,0.03);
    }
    .fmt-table td:first-child { color: var(--text); font-weight: 600; }
    .fmt-table code {
        font-family: 'JetBrains Mono', monospace; font-size: 0.7rem;
        color: #5a9ad8; background: var(--surface-0);
        padding: 0.1rem 0.35rem; border-radius: 3px;
    }

    /* ── EMPTY STATE ─────────────────────────────────────────── */
    .empty-state {
        text-align: center; padding: 3.5rem 1rem;
        border: 1px dashed var(--line);
        border-radius: var(--r);
        background: var(--surface-1);
        margin: 1rem 0;
    }
    .empty-state-icon {
        font-size: 2rem; margin-bottom: 0.9rem; line-height: 1;
        opacity: 0.4;
    }
    .empty-state-title {
        font-family: 'Barlow', sans-serif;
        font-size: 1rem; font-weight: 700; color: #d4dfe8; margin-bottom: 0.4rem;
    }
    .empty-state-body {
        font-size: 0.81rem; color: var(--text-3);
        max-width: 380px; margin: 0 auto; line-height: 1.65;
    }

    /* ── SPINNER ─────────────────────────────────────────────── */
    .stSpinner > div { border-top-color: var(--blue) !important; }

    /* ── SCROLLBAR ───────────────────────────────────────────── */
    ::-webkit-scrollbar { width: 5px; height: 5px; }
    ::-webkit-scrollbar-track { background: transparent; }
    ::-webkit-scrollbar-thumb { background: var(--line-2); border-radius: 3px; }
    ::-webkit-scrollbar-thumb:hover { background: rgba(255,255,255,0.16); }
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SHARED CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

NUM = r'\d+(?:[.,]\d+)?'
HIGH, MEDIUM, LOW = "high", "medium", "low"
MAX_REASONABLE_DAYS   = 365
MIN_TIME_WINDOW_SLOTS = 6

DISPLAY_FIELDS = ["User ID", "Level", "Construction", "Research", "Troops",
                  "FCs", "FC Shards", "Time UTC", "Days"]

FIELD_HINTS = {
    "Level":        "e.g. FC3, FC5",
    "Construction": "e.g. 24d 3h  or  42d  or  35",
    "Research":     "e.g. 42d  or  20",
    "Troops":       "e.g. 100d 10h  or  50",
    "FCs":          "Number of FCs, e.g. 2693 or 2700",
    "FC Shards":    "Number of shards, e.g. 434",
    "Time UTC":     "e.g. 16:00–19:00  or  14:30-17  or  13",
    "Days":         "e.g. Mon, Thu  or  1, 4",
}

FIELD_WARN_MSG = {
    LOW:    "⚠ Could not parse — please verify",
    MEDIUM: "⚠ Parsed with low confidence — please check",
}

SAMPLE_RAW = """\
User ID: 1
Level: FC1
CONSTRUCTION (Monday): 1
RESEARCH (Tuesday):  1
TROOPS (Thursday): 1
How many FCs and FC shards you have: FC 1 shards 1
Desired time UTC (minimum 3 hour window): 16.00-19.00
Desired day(s) (1,.2,.4): Mon, Tue

User ID: 2
Level: FC2
CONSTRUCTION (Monday): 2d 2h
RESEARCH (Tuesday): 2d 2m
TROOPS (Thursday): 2d 2h
How many FCs and FC shards you have: FC: 2, FC shards: 2  
Desired time UTC (minimum 3 hour window): 7utc till 21utc
Desired day(s): 2, 4

User ID: 3
Level: FC3
CONSTRUCTION (Monday): 3 days
RESEARCH (Tuesday): 3 day 1 hour
TROOPS (Thursday): 3 days
How many FCs and FC shards you have: 3 FCs, 3 FC shards
Desired time UTC (minimum 3 hour window): 00, 10, 11, 23
Desired day(s): Monday, Tuesday

User ID: 4
Level: FC4
CONSTRUCTION (Monday): 4d
RESEARCH (Tuesday): 4d 4h
TROOPS (Thursday): 4d
How many FCs and FC shards you have: 4 FC, 4 shards
Desired time UTC (minimum 3 hour window): 14-17
Desired day(s): 1, 2, 4

User ID: 5
Level: FC5
CONSTRUCTION (Monday): 7200min
RESEARCH (Tuesday): 5 days, 5 hours
TROOPS (Thursday): 5
How many FCs and FC shards you have: 5 Crystals, 5 shards
Desired time UTC (minimum 3 hour window): 00utc - 4utc, 9utc, 20-23
Desired day(s): Mon, Thu
"""

DAY_CONFIG = [
    {"day": 1, "label": "Day 1 — VP",  "col": "Construction"},
    {"day": 2, "label": "Day 2 — VP",  "col": "Research"},
    {"day": 4, "label": "Day 4 — MoE", "col": "Troops"},
]

# Day 4 VP is a virtual 5th config: re-runs Day 4 but only for unassigned players
DAY4_VP_CONFIG = {"day": 4, "label": "Day 4 — VP", "col": "Troops"}

# Extra per-day column shown in the timeline view
_DAY_EXTRA_COL = {1: "FCs", 2: "FC Shards", 4: None}


# ═══════════════════════════════════════════════════════════════════════════════
# UI HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def banner(kind: str, html: str) -> None:
    """Render an alert/info/success/error banner.

    Args:
        kind: One of 'alert', 'info', 'success', 'error'.
        html: Inner HTML content for the banner.
    """
    st.markdown(f'<div class="{kind}-banner">{html}</div>', unsafe_allow_html=True)


def stat_card(value, label: str, warn: bool = False) -> str:
    """Return HTML for a single stat card."""
    cls = "stat-card warn" if warn else "stat-card"
    return (
        f'<div class="{cls}">'
        f'<div class="stat-value">{value}</div>'
        f'<div class="stat-label">{label}</div>'
        f'</div>'
    )


def stat_row(*cards: str) -> None:
    """Render a row of stat cards."""
    st.markdown(
        f'<div class="stats-row">{"".join(cards)}</div>',
        unsafe_allow_html=True,
    )


def render_stepper(steps: list[tuple[str, str]]) -> None:
    """Render the workflow stepper.

    Args:
        steps: List of (label, state) tuples where state is 'done', 'active', or 'idle'.
    """
    parts = []
    for i, (label, state) in enumerate(steps):
        if i:
            parts.append('<span class="step-arrow">&#8250;</span>')
        parts.append(
            f'<div class="step">'
            f'<div class="step-circle {state}">{i + 1}</div>'
            f'<span class="step-label {state}">{label}</span>'
            f'</div>'
        )
    st.markdown(
        f'<div class="stepper">{"".join(parts)}</div>',
        unsafe_allow_html=True,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# PARSER HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_number_str(val: str) -> str:
    """Normalise a numeric string with thousand-separators to a plain decimal.

    Handles:
      "1.900"  → "1900"   (dot as thousands sep)
      "1,900"  → "1900"   (comma as thousands sep)
      "4.17"   → "4.17"   (dot as decimal — NOT stripped)
      "1,2"    → "1.2"    (comma as decimal sep → convert to dot)
    """
    # Dot-as-thousands: only when the fractional part is exactly 3 digits
    # AND there is no additional decimal point (e.g. "1.900" but not "4.175")
    dot_parts = val.split(".")
    if len(dot_parts) == 2 and len(dot_parts[-1]) == 3 and dot_parts[-1].isdigit() and dot_parts[0].isdigit():
        return val.replace(".", "")

    # Comma-as-thousands: same logic
    comma_parts = val.split(",")
    if len(comma_parts) == 2 and len(comma_parts[-1]) == 3 and comma_parts[-1].isdigit() and comma_parts[0].isdigit():
        return val.replace(",", "")

    # Comma-as-decimal separator (e.g. "1,2" or "4,5")
    return val.replace(",", ".")


def normalize_duration(raw: str) -> tuple:
    """Parse a speedup duration string into (days_float, confidence)."""
    if not raw:
        return "", LOW
    raw = raw.strip().lower()
    days = 0.0

    # Explicit range like "3d-5d" — take the lower bound
    range_m = re.match(rf'({NUM})\s*d?\s*[-\u2013]\s*({NUM})\s*d', raw)
    if range_m:
        val = float(range_m.group(1).replace(",", "."))
        conf = MEDIUM if val <= MAX_REASONABLE_DAYS else LOW
        return round(val, 2), conf

    day_m  = re.search(rf'({NUM})\s*(?:d|day|days)\b', raw)
    hour_m = re.search(rf'({NUM})\s*(?:h|hr|hour|hours)\b', raw)
    min_m  = re.search(rf'({NUM})\s*(?:m|min|minute|minutes)\b', raw)

    if day_m:  days += float(day_m.group(1).replace(",", "."))
    if hour_m: days += float(hour_m.group(1).replace(",", ".")) / 24
    if min_m:  days += float(min_m.group(1).replace(",", ".")) / 1440

    if days > 0:
        conf = MEDIUM if days > MAX_REASONABLE_DAYS else HIGH
        return round(days, 2), conf

    # Bare number fallback
    num_m = re.search(rf'({NUM})', raw)
    if num_m:
        val = float(num_m.group(1).replace(",", "."))
        conf = MEDIUM if val > MAX_REASONABLE_DAYS else HIGH
        return val, conf

    return "", LOW


def parse_fc_shards(raw: str) -> tuple:
    """Parse an FC / shards line into (fc_val, fc_conf, shard_val, shard_conf)."""
    if not raw:
        return "", LOW, "", LOW

    text = raw.strip().lower()
    fc_val = shard_val = None

    # 1. Update the regex to capture the punctuation (\d+(?:[.,]\d+)?)
    for pat in [r'shards?\s*[:\-]?\s*(\d+(?:[.,]\d+)?)', r'(\d+(?:[.,]\d+)?)\s*(?:fc\s+)?shards?']:
        m = re.search(pat, text, re.I)
        if m:
            try:
                # 2. Now m.group(1) will be "1.900", allowing _parse_number_str to clean it up!
                shard_val = int(float(_parse_number_str(m.group(1))))
            except ValueError:
                pass
            break

    for pat in [
        r'fcs?\s*[:\-]\s*(\d+(?:[.,]\d+)?)',
        r'fcs?\s+(\d+(?:[.,]\d+)?)',
        r'(\d+(?:[.,]\d+)?)\s*fcs?\b',
        r'(\d+(?:[.,]\d+)?)\s*crystals?',
    ]:
        m = re.search(pat, text, re.I)
        if m:
            try:
                fc_val = int(float(_parse_number_str(m.group(1))))
            except ValueError:
                pass
            break

    # Single bare number — treat as FC count with 0 shards
    if fc_val is None and shard_val is None:
        m = re.match(r'^\s*(\d+(?:[.,]\d+)?)\s*$', text)
        if m:
            try:
                fc_val    = int(float(_parse_number_str(m.group(1))))
                shard_val = 0
            except ValueError:
                pass

    return (
        fc_val    if fc_val    is not None else "",
        HIGH if fc_val    is not None else LOW,
        shard_val if shard_val is not None else "",
        HIGH if shard_val is not None else LOW,
    )


def parse_hhmm(s: str) -> int | None:
    """Parse 'HH:MM', 'HH.MM', or bare 'HH' into total minutes since midnight."""
    s = s.strip()
    m = re.match(r'^(\d{1,2})[.:](\d{2})$', s)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
    else:
        m = re.match(r'^(\d{1,2})$', s)
        if not m:
            return None
        h, mi = int(m.group(1)), 0
    return h * 60 + mi if (0 <= h <= 23 and 0 <= mi <= 59) else None


def normalize_time_utc(raw: str) -> tuple:
    """Parse an availability string into (slots_csv, confidence, slot_count)."""
    if not raw:
        return "", LOW, 0

    text = raw.lower().strip()
    text = text.replace("till", "-").replace(" to ", "-").replace("\u2013", "-")
    slots: set[int] = set()

    TIME_PAT  = r'\d{1,2}(?:[.:]\d{2})?'
    range_pat = rf'({TIME_PAT})\s*(?:utc)?\s*-\s*({TIME_PAT})\s*(?:utc)?'

    for m in re.finditer(range_pat, text):
        start = parse_hhmm(m.group(1))
        end   = parse_hhmm(m.group(2))
        if start is None or end is None:
            continue
        start = (start // 30) * 30
        end   = (end   // 30) * 30
        if start == end:
            continue
        # Walk 30-minute steps, wrapping midnight.
        # Guard against infinite loop: cross-midnight ranges are valid (e.g. 22:00-02:00)
        # but we must cap iterations to at most 48 steps (full 24 hours).
        t = start
        steps = 0
        while t != end and steps < 48:
            slots.add(t)
            t = (t + 30) % (24 * 60)
            steps += 1

    text_no_ranges = re.sub(range_pat, " ", text)
    for m in re.finditer(rf'\b({TIME_PAT})\s*(?:utc)?\b', text_no_ranges):
        t = parse_hhmm(m.group(1))
        if t is None:
            continue
        slot = (t // 30) * 30
        if 0 <= slot < 24 * 60:
            slots.add(slot)
            next_slot = slot + 30
            if next_slot < 24 * 60:
                slots.add(next_slot)

    if not slots:
        return "", LOW, 0

    valid     = sorted(s for s in slots if 0 <= s < 24 * 60)
    slots_str = ",".join(f"{s // 60:02d}:{s % 60:02d}" for s in valid)
    return slots_str, HIGH, len(valid)


def normalize_days(raw: str) -> tuple:
    """Parse a days string into (csv_of_day_numbers, confidence)."""
    if not raw:
        return "", LOW

    text = raw.lower()
    day_map = {
        "mon": "1", "monday": "1", "tue": "2", "tuesday": "2",
        "wed": "3", "wednesday": "3", "thu": "4", "thursday": "4",
        "fri": "5", "friday": "5", "sat": "6", "saturday": "6",
        "sun": "7", "sunday": "7",
    }
    found = list(re.findall(r'\b([1-7])\b', text))
    for k, v in day_map.items():
        if re.search(rf'\b{k}\b', text):
            found.append(v)
    found = sorted(set(found), key=int)
    return (",".join(found), HIGH) if found else ("", LOW)


def extract_field(block: str, patterns: list) -> str:
    """Return the first capture group from the first matching pattern."""
    for pat in patterns:
        m = re.search(pat, block, re.I | re.MULTILINE)
        if m:
            return m.group(1).strip()
    return ""


# Field-specific patterns (module-level to avoid repeated allocation)
_CONSTRUCTION_PATS = [
    r'CONSTRUCTION\s*(?:\([^)]*\))?\s*[:\-]\s*([^\n]+)',
    r'Constr[a-z]{2,10}\s*(?:\([^)]*\))?\s*[:\-]\s*([^\n]+)',
]
_RESEARCH_PATS = [
    r'RESEARCH\s*(?:\([^)]*\))?\s*[:\-]\s*([^\n]+)',
    r'Rese?[a-z]{2,7}\s*(?:\([^)]*\))?\s*[:\-]\s*([^\n]+)',
]
_TROOPS_PATS = [
    r'TROOPS?\s*(?:\([^)]*\))?\s*[:\-]\s*([^\n]+)',
    r'Troop[a-z]{0,4}\s*(?:\([^)]*\))?\s*[:\-]\s*([^\n]+)',
]

_SPEEDUP_FIELDS: list[tuple[str, list]] = [
    ("Construction", _CONSTRUCTION_PATS),
    ("Research",     _RESEARCH_PATS),
    ("Troops",       _TROOPS_PATS),
]

_FC_PATS = [
    r'(?:How many FCs[^:\n]*|FC[s]?\s*/\s*[Ss]hard[s]?[^:\n]*)\s*[:\-]?\s*([^\n]+)',
    r'FC[s]?\s+and[^:\n]*[:\-]?\s*([^\n]+)',
    r'(?:Crystal[s]?[^:\n]*)\s*[:\-]?\s*([^\n]+)',
]

_TIME_PATS = [
    r'Desired\s+time\s+UTC[^:\n]*[:\-]?\s*([^\n]+)',
    r'Time\s+UTC[^:\n]*[:\-]?\s*([^\n]+)',
    r'UTC\s*[:\-]\s*([^\n]+)',
]

_DAYS_PATS = [
    r'Desired\s+day(?:\(s\))?\s*(?:\([^)]*\))?\s*[:\-]\s*([^\n]+)',
    r'Day(?:s)?\s*(?:\([^)]*\))?\s*[:\-]\s*([^\n]+)',
]


def parse_block(block: str) -> dict:
    """Parse a single player block into a record dict with confidence metadata."""
    r: dict = {}

    r["User ID"] = extract_field(block, [
        r'User\s*ID\s*[:\-]?\s*(\d+)', r'\bID\s*[:\-]?\s*(\d+)',
    ])

    r["Level"]        = extract_field(block, [r'Level\s*[:\-]?\s*(\S+)', r'LVL\s*[:\-]?\s*(\S+)'])
    r["_conf_Level"]  = HIGH if r["Level"] else LOW

    for field, pats in _SPEEDUP_FIELDS:
        val, conf          = normalize_duration(extract_field(block, pats))
        r[field]           = val
        r[f"_conf_{field}"] = conf

    fc_line          = extract_field(block, _FC_PATS)
    r["_fc_raw"]     = fc_line
    fc_v, fc_c, sh_v, sh_c = parse_fc_shards(fc_line)
    r["FCs"]          = fc_v;  r["_conf_FCs"]       = fc_c
    r["FC Shards"]    = sh_v;  r["_conf_FC Shards"] = sh_c

    tv, tc, slot_count = normalize_time_utc(extract_field(block, _TIME_PATS).strip())
    r["Time UTC"] = tv
    if tv and slot_count < MIN_TIME_WINDOW_SLOTS:
        r["_conf_Time UTC"] = MEDIUM
        r["_warn_Time UTC"] = f"Only {slot_count / 2:.4g}h window — minimum is 3h"
    else:
        r["_conf_Time UTC"] = tc

    dv, dc     = normalize_days(extract_field(block, _DAYS_PATS))
    r["Days"]  = dv
    r["_conf_Days"] = dc

    return r


def parse_input(text: str) -> tuple[list, list, dict]:
    """Split raw text into player blocks and parse each one.

    Returns:
        (records, warnings, duplicates) where:
          - records   : list of dicts, one per unique User ID (first occurrence kept
                        until the caller resolves duplicates)
          - warnings  : list of human-readable issue strings (non-duplicate problems)
          - duplicates: {uid: [rec, rec, ...]} for every User ID that appeared > once
    """
    parts = re.split(r'(?=User\s*ID\s*[:\-]?\s*\d)', text.strip(), flags=re.I)
    # Drop any leading fragment that has no User ID (e.g. a blank preamble line)
    parts = [p.strip() for p in parts if p.strip() and re.search(r'User\s*ID\s*[:\-]?\s*\d', p, re.I)]

    all_parsed: list = []
    warnings:   list = []

    for i, part in enumerate(parts, 1):
        rec = parse_block(part)
        if not rec["User ID"]:
            warnings.append(f"Block {i}: Could not find User ID — skipped.")
            continue
        rec["_raw_block"] = part
        rec["_manual"]    = False
        all_parsed.append(rec)

    # Group by User ID
    by_uid: dict = {}
    for rec in all_parsed:
        by_uid.setdefault(rec["User ID"], []).append(rec)

    duplicates = {uid: recs for uid, recs in by_uid.items() if len(recs) > 1}
    # records = first occurrence for each uid (caller will replace resolved ones)
    records = [recs[0] for recs in by_uid.values()]

    return records, warnings, duplicates


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER (Parser export)
# ═══════════════════════════════════════════════════════════════════════════════

def build_excel(df_export: pd.DataFrame, flagged_cells: set) -> bytes:
    """Build a styled openpyxl workbook and return it as bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SvS Data"

    HEADER_FILL  = PatternFill("solid", fgColor="1A2535")
    FLAGGED_FILL = PatternFill("solid", fgColor="1E1608")
    EMPTY_FILL   = PatternFill("solid", fgColor="130E0E")
    THIN         = Side(style="thin", color="1E2533")
    BORDER       = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

    cols     = list(df_export.columns)
    uid_col  = cols.index("User ID") if "User ID" in cols else None

    # Header row
    for ci, col in enumerate(cols, 1):
        cell           = ws.cell(row=1, column=ci, value=col)
        cell.font      = Font(bold=True, color="7AABDC", size=10)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER
    ws.row_dimensions[1].height = 22

    # Data rows
    for ri, (_, row) in enumerate(df_export.iterrows(), 2):
        uid = str(row["User ID"]) if uid_col is not None else None
        for ci, col in enumerate(cols, 1):
            val  = row[col]
            cell = ws.cell(row=ri, column=ci, value=None if val in ("", "—") else val)
            cell.alignment = Alignment(vertical="center")
            cell.border    = BORDER
            if val in ("", "—", None):
                cell.fill = EMPTY_FILL
                cell.font = Font(color="4A5568")
            elif uid and (uid, col) in flagged_cells:
                cell.fill = FLAGGED_FILL
                cell.font = Font(color="D4A855")
            else:
                cell.font = Font(color="C9D1DC")
        ws.row_dimensions[ri].height = 18

    # Auto-fit column widths
    for ci, col in enumerate(cols, 1):
        cell_lengths = [len(str(v)) for v in df_export[col].fillna("")]
        max_len = max([len(str(col))] + cell_lengths, default=len(str(col)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 42)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════════════════
# SCHEDULER HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def slots_str_to_hours(slots_str: str) -> list[int]:
    """Convert parser Time UTC output ("14:00,14:30,15:00") to unique integer hours.

    Each "HH:MM" token contributes its hour component.  The result is a sorted,
    deduplicated list of integer hours (0-23) that the scheduler uses to build
    the set of available 30-minute slots via _user_slots().
    """
    if not slots_str:
        return []
    hours: set[int] = set()
    for tok in str(slots_str).split(","):
        tok = tok.strip()
        if ":" in tok:
            try:
                h = int(tok.split(":")[0])
                if 0 <= h <= 23:
                    hours.add(h)
            except ValueError:
                pass
        elif tok.isdigit():
            h = int(tok)
            if 0 <= h <= 23:
                hours.add(h)
    return sorted(hours)


def parse_ints(s) -> list[int]:
    """Parse a comma-separated string of integers, silently skipping bad tokens."""
    result = []
    for x in str(s).split(","):
        x = x.strip()
        if x:
            try:
                result.append(int(x))
            except ValueError:
                pass
    return result


def slot_label(slot: int) -> str:
    """Return a human-readable 30-minute window label, e.g. '14:00 – 14:30'.

    Args:
        slot: Integer index 0-47 (slot 0 = 00:00-00:30, slot 1 = 00:30-01:00, …).
    """
    start_min = slot * 30          # minutes since midnight
    end_min   = start_min + 30     # always 30 min later (may exceed 1440 for slot 47, but display wraps)
    sh, sm = divmod(start_min, 60)
    eh, em = divmod(end_min % (24 * 60), 60)
    return f"{sh:02d}:{sm:02d} – {eh:02d}:{em:02d}"


def _user_slots(user: dict) -> set[int]:
    """Return the set of 30-min slot indices for a user (h*2 and h*2+1 per hour)."""
    return {h * 2 + f for h in user["hours"] for f in (0, 1)}


# ─────────────────────────────────────────────────────────────────────────────
# MIN-COST MAX-FLOW SCHEDULER
#
# Objective (strict priority order):
#   1. Maximise the number of users assigned a slot  (max cardinality)
#   2. Among all maximum-cardinality solutions, maximise total score
#      — i.e. the highest-scoring users always win when slots are contested.
#
# Why MCMF instead of Hopcroft-Karp + swap pass?
# ────────────────────────────────────────────────
# The old HK approach ran BFS augmentation for cardinality, then tried a
# one-level score-swap pass to improve fairness.  That swap pass was too
# shallow: it could only displace one user at a time, so any inversion
# requiring a chain of re-routings was left unsolved.
#
# MCMF solves both objectives simultaneously in a single pass:
#
#   Graph:  S → user_i  (cap 1, cost 0)
#           user_i → slot_j  (cap 1, cost ∝ max_score − score_i)
#           slot_j → T  (cap 1, cost 0)
#
#   networkx.max_flow_min_cost() pushes the maximum feasible flow first
#   (= max users placed), then, among all maximum flows, minimises total
#   edge cost (= maximises total score of placed users).
#
# This is provably optimal: no assignment can simultaneously place more
# users AND rank higher-scorers better.
#
# Remaining "inversions" (an unplaced user scores higher than the lowest
# placed user) are mathematically unavoidable.  They only occur when the
# unplaced user's entire time window is saturated by higher-scorers who
# ALSO need those same slots, and no rerouting frees a slot without
# dropping someone — MCMF already tried every possible rerouting.
# ─────────────────────────────────────────────────────────────────────────────

def run_day(users: list, dc: dict) -> dict:
    """Run the MCMF scheduler for a single day config and return results dict."""
    col = dc["col"]
    day = dc["day"]

    # ── 1. Deduplicate by User ID ──────────────────────────────────────────────
    # Keep the highest-scoring entry when the same User ID appears twice.
    best_by_uid: dict = {}
    for u in users:
        uid = u["User ID"]
        if uid not in best_by_uid or u[col] > best_by_uid[uid][col]:
            best_by_uid[uid] = u

    eligible = [u for u in best_by_uid.values() if day in u["days"] and u["hours"]]
    if not eligible:
        return {
            "day": day, "col": col,
            "slot_occ": {}, "user_slot": {},
            "via_reshuffle": set(), "moved": {}, "chains": [],
            "unassigned": [], "eligible": [],
        }

    eligible = sorted(eligible, key=lambda u: u[col], reverse=True)

    # ── 2. Apply manual slot overrides ────────────────────────────────────────
    # Pinned users are pre-placed; the solver only runs on the rest.
    pinned_slot_occ:  dict = {}   # slot → uid
    pinned_user_slot: dict = {}   # uid  → slot
    free_eligible:    list = []

    for u in eligible:
        uid     = u["User ID"]
        uid_str = str(uid)
        override_slot = u.get("_slot_overrides", {}).get(day)
        if override_slot is not None and 0 <= override_slot < 48:
            # If two pins claim the same slot the later one (lower score) is ignored
            if override_slot not in pinned_slot_occ:
                pinned_slot_occ[override_slot]  = uid
                pinned_user_slot[uid]           = override_slot
            else:
                free_eligible.append(u)   # pin conflict — fall through to solver
        else:
            free_eligible.append(u)

    # ── 3. Build slot universe (from all eligible, including pinned) ───────────
    all_slots = sorted({h * 2 + f for u in eligible for h in u["hours"] for f in (0, 1)})
    # Also include pinned slots even if they fall outside normal availability
    for s in pinned_slot_occ:
        if s not in all_slots:
            all_slots = sorted(set(all_slots) | {s})
    slot_set = set(all_slots)

    # ── 4. Edge costs (for free_eligible only) ─────────────────────────────────
    SCALE  = 10_000
    max_s  = eligible[0][col] if eligible else 1.0
    free_costs = [int((max_s - u[col]) * SCALE) + i for i, u in enumerate(free_eligible)]

    # ── 5. Build flow network (excluding already-pinned slots) ─────────────────
    available_slots = sorted(s for s in all_slots if s not in pinned_slot_occ)

    slot_occ:  dict = dict(pinned_slot_occ)
    user_slot: dict = dict(pinned_user_slot)

    if free_eligible and available_slots:
        S, T       = "S", "T"
        user_nodes = [f"U{i}" for i in range(len(free_eligible))]
        slot_node  = {s: f"SL{s}" for s in available_slots}

        G = nx.DiGraph()
        for i in range(len(free_eligible)):
            G.add_edge(S, user_nodes[i], capacity=1, weight=0)

        for i, u in enumerate(free_eligible):
            u_slots = _user_slots(u) & set(available_slots)
            for s in u_slots:
                G.add_edge(user_nodes[i], slot_node[s], capacity=1, weight=free_costs[i])

        for s in available_slots:
            G.add_edge(slot_node[s], T, capacity=1, weight=0)

        # ── 6. Solve ───────────────────────────────────────────────────────────
        flow_dict = nx.max_flow_min_cost(G, S, T)

        # ── 7. Extract assignments ─────────────────────────────────────────────
        for i, u in enumerate(free_eligible):
            uid = u["User ID"]
            un  = user_nodes[i]
            for s in available_slots:
                if flow_dict.get(un, {}).get(slot_node[s], 0) == 1:
                    slot_occ[s]    = uid
                    user_slot[uid] = s
                    break

    # ── 8. Build unassigned list with blocker detail ───────────────────────────
    # Find the minimum score among all assigned users (or +inf if nobody was placed).
    min_assigned_score = min(
        (u[col] for u in eligible if u["User ID"] in user_slot),
        default=float("inf"),
    )

    unassigned = []
    for u in eligible:
        uid = u["User ID"]
        if uid in user_slot:
            continue
        all_u_slots = _user_slots(u)
        blockers    = {slot_occ[s] for s in all_u_slots if s in slot_occ}
        names       = [f"{b}" for b in sorted(str(b) for b in blockers)]

        if u[col] < min_assigned_score:
            unassigned.append({
                "user":   u,
                "reason": "not enough speedups",
                "detail": f"Score {u[col]:.2f} < minimum placed {min_assigned_score:.2f}.",
            })
        else:
            unassigned.append({
                "user":   u,
                "reason": "Window saturated",
                "detail": (
                    f"All {len(all_u_slots)} slot(s) in their time window are taken — "
                    f"Blocked by: {', '.join(names) or 'unknown'}."
                ),
            })

    return {
        "day":           day,
        "col":           col,
        "slot_occ":      slot_occ,
        "user_slot":     user_slot,
        "via_reshuffle": set(),
        "moved":         {},
        "chains":        [],
        "unassigned":    unassigned,
        "eligible":      eligible,
    }


def run_scheduler(users: list) -> list:
    """Run the scheduler for all DAY_CONFIG entries and return a list of results."""
    return [run_day(users, dc) for dc in DAY_CONFIG]


def run_day4_vp(users: list, day4_result: dict) -> dict:
    """Run a second VP scheduling pass for Day 4 using only unassigned players.

    Players who were unassigned in the Day 4 MoE run (window saturated or low
    score) get a second chance here.  Scheduling logic is identical to run_day
    (MCMF over Troops score) but the eligible pool is restricted to those
    unassigned IDs only.
    """
    unassigned_ids = {e["user"]["User ID"] for e in day4_result["unassigned"]}
    if not unassigned_ids:
        return {
            "day": 4, "col": "Troops",
            "slot_occ": {}, "user_slot": {},
            "via_reshuffle": set(), "moved": {}, "chains": [],
            "unassigned": [], "eligible": [],
            "is_vp4": True,
        }

    # Restrict to users who were unassigned on Day 4 and have time windows
    vp4_users = [u for u in users if u["User ID"] in unassigned_ids and 4 in u["days"]]

    result = run_day(vp4_users, DAY4_VP_CONFIG)
    result["is_vp4"] = True
    return result


# ─────────────────────────────────────────────────────────────────────────────
# DataFrame builders
# ─────────────────────────────────────────────────────────────────────────────

def build_timeline_df(users: list, dr: dict, extra_col: str | None = None) -> pd.DataFrame:
    """Build the slot-timeline DataFrame for a single day result."""
    col        = dr["col"]
    avail_slots = sorted({
        h * 2 + f
        for u in dr["eligible"]
        for h in u["hours"]
        for f in (0, 1)
    })
    # Build a lookup for quick user access
    user_by_id = {u["User ID"]: u for u in users}

    rows = []
    for slot in avail_slots:
        auid = dr["slot_occ"].get(slot)
        au   = user_by_id.get(auid)
        if au:
            assigned_id    = f"{au['User ID']}"
            assigned_score = au[col]
        else:
            assigned_id, assigned_score = "empty", None

        row = {
            "Time Slot": slot_label(slot),
            "Assigned":  assigned_id,
            "Speedups":  f"{assigned_score:.2f}" if assigned_score is not None else "—",
        }
        if extra_col:
            row[extra_col] = (
                str(int(au[extra_col]))
                if au and au.get(extra_col) not in (None, "", "—")
                else "—"
            )
        rows.append(row)
    return pd.DataFrame(rows)


def build_unassigned_df(dr: dict) -> pd.DataFrame:
    """Build the unassigned-users DataFrame for a single day result."""
    return pd.DataFrame([{
        "User ID":  f"{e['user']['User ID']}",
        "Speedups": f"{e['user'][dr['col']]:.2f}",
        "Reason":   e.get("reason", ""),
        "Detail":   e.get("detail", "—"),
    } for e in dr["unassigned"]])


def build_summary_df(users: list, day_results: list, vp4_result: dict | None = None) -> pd.DataFrame:
    """Build the cross-day summary DataFrame."""
    day_map = {dr["day"]: dr for dr in day_results}

    # Pre-build sets for O(1) lookups inside the loop
    day4_dr = day_map.get(4)
    day4_unassigned_ids: set = (
        {e["user"]["User ID"] for e in day4_dr["unassigned"]} if day4_dr else set()
    )

    rows = []
    for u in users:
        row = {"User ID": u["User ID"], "Level": u["Level"]}
        for dc in DAY_CONFIG:
            dr  = day_map[dc["day"]]
            col = dc["col"]
            if dc["day"] not in u["days"]:
                row[dc["label"]] = "—"
            elif u["User ID"] in dr["user_slot"]:
                slot = dr["user_slot"][u["User ID"]]
                row[dc["label"]] = f"{slot_label(slot)}  [{u[col]:.2f}]"
            else:
                row[dc["label"]] = f"❌ unassigned  [{u[col]:.2f}]"

        # Day 4 — VP column: only meaningful for players unassigned in Day 4 MoE
        if vp4_result is not None:
            col = "Troops"
            if 4 not in u["days"]:
                row[DAY4_VP_CONFIG["label"]] = "—"
            elif day4_dr and u["User ID"] in day4_dr["user_slot"]:
                # Already placed in MoE — not eligible for VP
                row[DAY4_VP_CONFIG["label"]] = "— (MoE placed)"
            elif u["User ID"] in vp4_result["user_slot"]:
                slot = vp4_result["user_slot"][u["User ID"]]
                row[DAY4_VP_CONFIG["label"]] = f"{slot_label(slot)}  [{u[col]:.2f}]"
            elif u["User ID"] in day4_unassigned_ids:
                row[DAY4_VP_CONFIG["label"]] = f"❌ unassigned  [{u[col]:.2f}]"
            else:
                row[DAY4_VP_CONFIG["label"]] = "—"

        rows.append(row)
    return pd.DataFrame(rows)


def to_excel_schedule(users: list, day_results: list, vp4_result: dict | None = None) -> io.BytesIO:
    """Export scheduler results to a multi-sheet xlsx file."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        build_summary_df(users, day_results, vp4_result).to_excel(
            writer, sheet_name="Summary", index=False
        )
        for dr in day_results:
            dc    = next(d for d in DAY_CONFIG if d["day"] == dr["day"])
            sheet = dc["label"].replace(" — ", " ").replace(" ", "_")[:31]
            build_timeline_df(users, dr).to_excel(
                writer, sheet_name=sheet, index=False
            )
            ua = build_unassigned_df(dr)
            if not ua.empty:
                ua.to_excel(
                    writer, sheet_name=f"Unassigned_Day{dc['day']}", index=False
                )
        # Day 4 — VP sheet
        if vp4_result is not None and (vp4_result["user_slot"] or vp4_result["unassigned"]):
            build_timeline_df(users, vp4_result).to_excel(
                writer, sheet_name="Day_4_VP", index=False
            )
            ua_vp4 = build_unassigned_df(vp4_result)
            if not ua_vp4.empty:
                ua_vp4.to_excel(writer, sheet_name="Unassigned_Day4_VP", index=False)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════════
# SESSION STATE & BROWSER STORAGE INIT
# ═══════════════════════════════════════════════════════════════════════════════

from streamlit_local_storage import LocalStorage

local_storage = LocalStorage()

browser_saved_data = local_storage.getItem("svs_raw_input")
initial_text       = browser_saved_data if browser_saved_data and browser_saved_data.strip() else SAMPLE_RAW

_SESSION_DEFAULTS = {
    "page":              "parser",
    "raw_input":         initial_text,
    "manual_records":    [],
    "excluded_ids":      set(),
    "corrections":       {},
    "_last_hash":        None,
    "_clipboard_csv":    None,
    "parsed_df":         None,
    # Duplicate resolution: {uid: index} — which occurrence to keep (0-based)
    "dup_choices":       {},
    # Manual slot overrides for the scheduler: {uid: {day: slot_index or None}}
    "slot_overrides":    {},
    # Run history for changelog: list of {label, summary_df, user_slot_map}
    "run_history":       [],
}
for _k, _v in _SESSION_DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v


# ═══════════════════════════════════════════════════════════════════════════════
# TOP NAVIGATION BAR
# ═══════════════════════════════════════════════════════════════════════════════

if st.session_state["parsed_df"] is not None:
    n           = len(st.session_state["parsed_df"])
    badge_html  = (
        f'<span class="topnav-badge">&#10003;&nbsp;{n} player{"s" if n != 1 else ""} loaded</span>'
    )
else:
    badge_html  = (
        '<span class="topnav-badge dim">awaiting data</span>'
    )

st.markdown(f"""
<div class="topnav">
  <div class="topnav-inner">
    <div class="topnav-brand">
      <div class="topnav-icon">
        <svg viewBox="0 0 18 18" fill="none" xmlns="http://www.w3.org/2000/svg">
          <defs>
            <linearGradient id="sg" x1="9" y1="1" x2="9" y2="14" gradientUnits="userSpaceOnUse">
              <stop offset="0%" stop-color="#e8cc80"/>
              <stop offset="100%" stop-color="#9a7230"/>
            </linearGradient>
          </defs>
          <path d="M9 1 L9.8 11.5 L9 14 L8.2 11.5 Z" fill="url(#sg)"/>
          <rect x="5" y="11" width="8" height="1.5" rx="0.75" fill="#b89040"/>
          <rect x="7.5" y="12.5" width="3" height="4" rx="1" fill="#7a5820"/>
          <circle cx="9" cy="17" r="1" fill="#9a7230"/>
        </svg>
      </div>
      <div class="topnav-text">
        <div class="topnav-name">SvS <em>#3442</em></div>
        <div class="topnav-tagline">Ministry Scheduler</div>
      </div>
    </div>
    <div class="topnav-right">
      <span class="topnav-meta">v2.0</span>
      <div class="topnav-sep"></div>
      {badge_html}
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

nav_col1, nav_col2, _ = st.columns([1.5, 1.5, 5])
with nav_col1:
    if st.button(
        "Parser", use_container_width=True,
        type="primary" if st.session_state["page"] == "parser" else "secondary",
    ):
        st.session_state["page"] = "parser"
        st.rerun()
with nav_col2:
    if st.button(
        "Scheduler", use_container_width=True,
        type="primary" if st.session_state["page"] == "scheduler" else "secondary",
    ):
        st.session_state["page"] = "scheduler"
        st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE 1 — PARSER
# ═══════════════════════════════════════════════════════════════════════════════

if st.session_state["page"] == "parser":

    has_data = st.session_state["parsed_df"] is not None
    render_stepper([
        ("Paste player data",  "active"),
        ("Review & correct",   "active"),
        ("Send to Scheduler",  "done" if has_data else "idle"),
        ("Run &amp; export",   "idle"),
    ])

    with st.expander("How to use this tool", expanded=(not has_data)):
        st.markdown("""
        <div class="guide-grid">
          <div class="guide-card">
            <div class="guide-card-icon">Step 01</div>
            <div class="guide-card-title">Paste player responses</div>
            <div class="guide-card-body">
              Copy the raw sign-up replies from WOS
              and paste the whole block into the <b>Raw Input</b> box. Each player entry
              must start with <code>User ID: &lt;number&gt;</code>.
            </div>
          </div>
          <div class="guide-card">
            <div class="guide-card-icon">Step 02</div>
            <div class="guide-card-title">Fix flagged fields</div>
            <div class="guide-card-body">
              Fields the parser couldn't read confidently are highlighted in amber below.
              Click the field and type the correct value. You can also add missing players
              with the <b>Add player manually</b> form.
            </div>
          </div>
          <div class="guide-card">
            <div class="guide-card-icon">Step 03</div>
            <div class="guide-card-title">Send to Scheduler</div>
            <div class="guide-card-body">
              Once everything looks right, click <b>Send to Scheduler →</b> at the bottom.
              Switch to the <b>Scheduler</b> tab, then hit <b>Run scheduler</b> to assign
              every player a 30-minute slot on each of their days.
            </div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <table class="fmt-table">
          <tr><th>Field</th><th>What it means</th><th>Accepted formats</th></tr>
          <tr><td>User ID</td><td>Unique player number</td><td><code>1</code>, <code>12345</code></td></tr>
          <tr><td>Level</td><td>FC tier</td><td><code>FC1</code> … <code>FC5</code></td></tr>
          <tr><td>Construction / Research / Troops</td><td>Speedups in days</td>
              <td><code>4d</code>, <code>3d 12h</code>, <code>7200min</code>, <code>5</code></td></tr>
          <tr><td>FCs / FC Shards</td><td>Resource counts</td>
              <td><code>FC 2693 shards 434</code>, <code>2,700 FCs</code></td></tr>
          <tr><td>Time UTC</td><td>Available hours (min 3 h window)</td>
              <td><code>14-17</code>, <code>16:00-19:30</code>, <code>00utc - 4utc, 20-23</code></td></tr>
          <tr><td>Days</td><td>Which SvS days the player joins</td>
              <td><code>1, 2, 4</code>, <code>Mon, Thu</code>, <code>Monday</code></td></tr>
        </table>
        <div style="font-size:0.73rem;color:#3a4558;margin-top:0.6rem;">
          Days: 1 = Construction &nbsp;·&nbsp; 2 = Research &nbsp;·&nbsp; 4 = Troops
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<div style='margin-bottom:1rem'></div>", unsafe_allow_html=True)

    col_input, col_tools = st.columns([4, 1])

    def _save_to_browser():
        local_storage.setItem("svs_raw_input", st.session_state["raw_input"])

    with col_input:
        st.markdown('<div class="section-label">Raw Input</div>', unsafe_allow_html=True)
        raw_text = st.text_area(
            "raw",
            height=340,
            label_visibility="collapsed",
            key="raw_input",
            on_change=_save_to_browser,
        )

    with col_tools:
        st.markdown('<div class="section-label">Input File</div>', unsafe_allow_html=True)
        uploaded = st.file_uploader("Load .txt", type=["txt"], label_visibility="collapsed",
                                    accept_multiple_files=True)
        if uploaded:
            merged = "\n".join(f.read().decode("utf-8") for f in uploaded)
            st.session_state["raw_input"] = merged
            st.rerun()
        st.download_button(
            "💾 Save raw input",
            data=st.session_state["raw_input"].encode("utf-8"),
            file_name="svs_input.txt",
            mime="text/plain",
            use_container_width=True,
        )

    if not raw_text.strip():
        st.info("Paste your player data in the text box above.")
        st.stop()

    records, parse_warnings, duplicates = parse_input(raw_text)
    input_hash = hash(raw_text)

    # ── Reset corrections when input changes ──────────────────────────────────
    if st.session_state["_last_hash"] != input_hash:
        new_corr = {
            rec["User ID"]: {f: rec[f] for f in DISPLAY_FIELDS if f != "User ID"}
            for rec in records
        }
        for mrec in st.session_state["manual_records"]:
            uid = mrec["User ID"]
            if uid not in new_corr:
                new_corr[uid] = {f: mrec.get(f, "") for f in DISPLAY_FIELDS if f != "User ID"}
        st.session_state["corrections"] = new_corr
        st.session_state["_last_hash"]  = input_hash
        # Clear stale dup choices for UIDs no longer duplicated
        st.session_state["dup_choices"] = {
            uid: v for uid, v in st.session_state["dup_choices"].items()
            if uid in duplicates
        }

    # ── Duplicate resolution UI ───────────────────────────────────────────────
    if duplicates:
        banner(
            "alert",
            f"<b>{len(duplicates)} duplicate User ID{'s' if len(duplicates) > 1 else ''} found.</b> "
            "Choose which entry to keep for each.",
        )
        for uid, dup_recs in duplicates.items():
            with st.expander(f"⚠ Duplicate — User ID {uid}  ({len(dup_recs)} entries)", expanded=True):
                option_labels = []
                for idx, rec in enumerate(dup_recs):
                    parts_preview = ", ".join(
                        f"{f}: {rec.get(f, '—')}"
                        for f in ("Construction", "Research", "Troops", "Time UTC", "Days")
                        if rec.get(f)
                    )
                    option_labels.append(f"Entry {idx + 1}  —  {parts_preview}")

                current_choice = st.session_state["dup_choices"].get(uid, 0)
                chosen = st.radio(
                    "Keep which entry?",
                    options=list(range(len(dup_recs))),
                    format_func=lambda i, _labels=option_labels: _labels[i],
                    index=current_choice,
                    key=f"dup_{uid}",
                    horizontal=False,
                )
                st.session_state["dup_choices"][uid] = chosen

        # Apply choices: replace the first-occurrence record with the chosen one
        choice_map = st.session_state["dup_choices"]
        for i, rec in enumerate(records):
            uid = rec["User ID"]
            if uid in duplicates:
                chosen_idx = choice_map.get(uid, 0)
                chosen_rec = duplicates[uid][chosen_idx]
                
                records[i] = chosen_rec

                st.session_state["corrections"][uid] = {
                    f: chosen_rec[f]
                    for f in DISPLAY_FIELDS
                    if f != "User ID"
                }

    for mrec in st.session_state["manual_records"]:
        uid = mrec["User ID"]
        if uid not in st.session_state["corrections"]:
            st.session_state["corrections"][uid] = {
                f: mrec.get(f, "") for f in DISPLAY_FIELDS if f != "User ID"
            }

    for w in parse_warnings:
        banner("error", f"⚠ {w}")

    all_records     = records + st.session_state["manual_records"]
    visible_records = [
        r for r in all_records
        if r["User ID"] not in st.session_state["excluded_ids"]
    ]

    if not all_records:
        st.warning("No valid records found. Make sure each block starts with 'User ID: <number>'.")
        st.stop()

    uncertain = [
        rec for rec in visible_records
        if any(
            rec.get(f"_conf_{f}", HIGH) in (LOW, MEDIUM)
            for f in DISPLAY_FIELDS if f != "User ID"
        )
    ]
    n_flagged = len(uncertain)

    stat_row(
        stat_card(len(visible_records), "Players"),
        stat_card(n_flagged,            "Need Review",   warn=bool(n_flagged)),
        stat_card(len(duplicates),      "Duplicates",    warn=bool(duplicates)),
        stat_card(len(st.session_state["excluded_ids"]), "Excluded"),
    )

    if uncertain:
        banner(
            "alert",
            f"<b>{n_flagged} record{'s' if n_flagged > 1 else ''} require attention</b> — "
            "some fields could not be parsed confidently. Review and correct below before exporting.",
        )

        for rec in uncertain:
            uid     = rec["User ID"]
            flagged = [
                f for f in DISPLAY_FIELDS
                if f != "User ID" and rec.get(f"_conf_{f}", HIGH) in (LOW, MEDIUM)
            ]
            label = f"User {uid}  ·  {len(flagged)} field{'s' if len(flagged) > 1 else ''} flagged"
            if rec.get("_manual"):
                label += "  [manual]"

            with st.expander(label, expanded=True):
                for f in flagged:
                    if rec.get(f"_warn_{f}"):
                        banner("info", f"ℹ {rec[f'_warn_{f}']}")

                if "_fc_raw" in rec and any(f in flagged for f in ("FCs", "FC Shards")):
                    st.markdown(
                        f'<div class="fc-raw">FC line: <span>{rec["_fc_raw"]}</span></div>',
                        unsafe_allow_html=True,
                    )

                n_cols  = min(len(flagged), 3)
                cols_w  = st.columns(n_cols)
                for i, field in enumerate(flagged):
                    conf = rec.get(f"_conf_{field}", HIGH)
                    with cols_w[i % n_cols]:
                        new_val = st.text_input(
                            label=field,
                            value=st.session_state["corrections"].get(uid, {}).get(field, ""),
                            key=f"fix_{uid}_{field}",
                            placeholder=FIELD_HINTS.get(field, ""),
                        )
                        st.markdown(
                            f'<div class="field-warn">{FIELD_WARN_MSG.get(conf, "")}</div>',
                            unsafe_allow_html=True,
                        )
                        st.session_state["corrections"].setdefault(uid, {})[field] = new_val

    with st.expander("➕ Add player manually"):
        with st.form("manual_add_form", clear_on_submit=True):
            c1, c2, c3 = st.columns(3)
            with c1:
                m_uid    = st.text_input("User ID *",    placeholder="e.g. 99999")
                m_level  = st.text_input("Level",        placeholder="e.g. FC3")
                m_constr = st.text_input("Construction", placeholder="e.g. 24d 3h")
            with c2:
                m_res    = st.text_input("Research",     placeholder="e.g. 42d")
                m_troops = st.text_input("Troops",       placeholder="e.g. 5d")
                m_fcs    = st.text_input("FCs",          placeholder="e.g. 2700")
            with c3:
                m_shards = st.text_input("FC Shards",    placeholder="e.g. 434")
                m_time   = st.text_input("Time UTC",     placeholder="e.g. 14:00-17:30")
                m_days   = st.text_input("Days",         placeholder="e.g. 1, 2")
            submitted = st.form_submit_button("Add Player", use_container_width=True)

        if submitted:
            uid_str = m_uid.strip()
            if not uid_str:
                st.error("User ID is required.")
            elif uid_str in [r["User ID"] for r in all_records]:
                st.error(f"User ID {uid_str} already exists.")
            else:
                constr_v, constr_c      = normalize_duration(m_constr)
                res_v,    res_c         = normalize_duration(m_res)
                troops_v, troops_c      = normalize_duration(m_troops)
                fc_raw                  = f"FC {m_fcs} shards {m_shards}" if (m_fcs or m_shards) else ""
                fc_v, fc_c, sh_v, sh_c = parse_fc_shards(fc_raw)
                time_v, time_c, scount  = normalize_time_utc(m_time)
                if time_v and scount < MIN_TIME_WINDOW_SLOTS:
                    time_c = MEDIUM
                days_v, days_c = normalize_days(m_days)

                new_rec = {
                    "User ID": uid_str,
                    "Level":        m_level.strip(),  "_conf_Level":        HIGH if m_level.strip() else LOW,
                    "Construction": constr_v,          "_conf_Construction": constr_c,
                    "Research":     res_v,             "_conf_Research":     res_c,
                    "Troops":       troops_v,          "_conf_Troops":       troops_c,
                    "FCs":          fc_v,              "_conf_FCs":          fc_c,
                    "FC Shards":    sh_v,              "_conf_FC Shards":    sh_c,
                    "Time UTC":     time_v,            "_conf_Time UTC":     time_c,
                    "Days":         days_v,            "_conf_Days":         days_c,
                    "_fc_raw":      fc_raw,
                    "_raw_block":   "(manual entry)",
                    "_manual":      True,
                }
                if time_v and scount < MIN_TIME_WINDOW_SLOTS:
                    new_rec["_warn_Time UTC"] = f"Only {scount / 2:.4g}h window — minimum is 3h"

                st.session_state["manual_records"].append(new_rec)
                st.session_state["corrections"][uid_str] = {
                    f: new_rec[f] for f in DISPLAY_FIELDS if f != "User ID"
                }
                st.success(f"Player {uid_str} added.")
                st.rerun()

    with st.expander("Manage records"):
        all_uids       = [r["User ID"] for r in all_records]
        manual_uid_set = {r["User ID"] for r in st.session_state["manual_records"]}

        col_del, col_clear = st.columns([3, 1])
        with col_del:
            to_exclude = st.multiselect(
                "Exclude from results and export",
                options=all_uids,
                default=[uid for uid in st.session_state["excluded_ids"] if uid in all_uids],
                format_func=lambda uid: f"User {uid}" + (" [manual]" if uid in manual_uid_set else ""),
            )
            st.session_state["excluded_ids"] = set(to_exclude)
        with col_clear:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Clear all manual", use_container_width=True):
                manual_ids = {r["User ID"] for r in st.session_state["manual_records"]}
                st.session_state["manual_records"] = []
                st.session_state["excluded_ids"]  -= manual_ids
                st.rerun()

        if st.session_state["excluded_ids"]:
            banner(
                "info",
                f"ℹ {len(st.session_state['excluded_ids'])} record(s) excluded from results and export.",
            )

    # ── Build the final DataFrame ──────────────────────────────────────────────
    final: list      = []
    conf_lookup: dict = {}

    for rec in visible_records:
        uid = rec["User ID"]
        row = {f: rec[f] for f in DISPLAY_FIELDS}
        row.update(st.session_state["corrections"].get(uid, {}))
        final.append(row)
        for f in DISPLAY_FIELDS:
            conf_lookup[(uid, f)] = rec.get(f"_conf_{f}", HIGH)

    df         = pd.DataFrame(final, columns=DISPLAY_FIELDS)
    df_display = df.replace("", "—")
    df_display[""] = [
        "✎" if any(r["User ID"] == uid and r.get("_manual") for r in all_records) else ""
        for uid in df_display["User ID"]
    ]
    df_display["_Raw Input"] = [
        next((rec["_raw_block"] for rec in all_records if rec["User ID"] == uid), "")
        for uid in df_display["User ID"]
    ]

    st.markdown('<div class="section-label" style="margin-top:0.5rem">Results</div>',
                unsafe_allow_html=True)
    st.dataframe(
        df_display,
        use_container_width=True,
        hide_index=True,
        column_config={
            "User ID":      st.column_config.TextColumn("User ID",       width="medium"),
            "Level":        st.column_config.TextColumn("Level",         width="small"),
            "Construction": st.column_config.NumberColumn("Construction", format="%.2f", width="small"),
            "Research":     st.column_config.NumberColumn("Research",     format="%.2f", width="small"),
            "Troops":       st.column_config.NumberColumn("Troops",       format="%.2f", width="small"),
            "FCs":          st.column_config.NumberColumn("FCs",          format="%d",   width="small"),
            "FC Shards":    st.column_config.NumberColumn("Shards",       format="%d",   width="small"),
            "Time UTC":     st.column_config.TextColumn("Time (UTC)",    width="large"),
            "Days":         st.column_config.TextColumn("Days",          width="small"),
            "":             st.column_config.TextColumn("",              width="small",
                                help="✎ = manually added record"),
            "_Raw Input":   st.column_config.TextColumn("Original Input",
                                help="Hover to see raw player input", width="medium"),
        },
    )

    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="section-label">Export &amp; send</div>', unsafe_allow_html=True)

    if st.button("🗓️  Send to Scheduler →", use_container_width=True, type="primary"):
        st.session_state["parsed_df"] = df.copy()
        st.session_state["page"]      = "scheduler"
        st.rerun()

    now           = datetime.utcnow().strftime("%Y%m%d_%H%M")
    df_export     = df[DISPLAY_FIELDS].copy()
    flagged_cells = {
        (uid, field)
        for (uid, field), conf in conf_lookup.items()
        if conf in (LOW, MEDIUM)
    }

    col_csv, col_xlsx, col_clip = st.columns(3)
    with col_csv:
        st.download_button(
            "⬇ CSV",
            data=df_export.to_csv(index=False).encode("utf-8"),
            file_name=f"svs_3442_{now}.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with col_xlsx:
        if HAS_OPENPYXL:
            st.download_button(
                "⬇ Excel",
                data=build_excel(df_export, flagged_cells),
                file_name=f"svs_3442_{now}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            banner("info", "Install <code>openpyxl</code> to enable Excel export.")
    with col_clip:
        if st.button("📋 Copy CSV", use_container_width=True):
            st.session_state["_clipboard_csv"] = df_export.to_csv(index=False)

    if st.session_state["_clipboard_csv"]:
        st.markdown(
            '<div class="section-label" style="margin-top:0.75rem">CSV — click icon to copy</div>',
            unsafe_allow_html=True,
        )
        st.code(st.session_state["_clipboard_csv"], language=None)


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE 2 — SCHEDULER
# ═══════════════════════════════════════════════════════════════════════════════

elif st.session_state["page"] == "scheduler":

    has_data = st.session_state["parsed_df"] is not None
    render_stepper([
        ("Paste player data",  "done"),
        ("Review &amp; correct", "done"),
        ("Send to Scheduler",  "done" if has_data else "idle"),
        ("Run &amp; export",   "active"),
    ])

    if not has_data:
        st.markdown("""
        <div class="empty-state">
          <div class="empty-state-icon">&#128197;</div>
          <div class="empty-state-title">No player data yet</div>
          <div class="empty-state-body">
            Go to the <b>Parser</b> page, paste your player responses, fix any flagged fields,
            then click <b>Send to Scheduler &#8594;</b> to load data here.
          </div>
        </div>
        """, unsafe_allow_html=True)

    SCHEDULER_SAMPLE = pd.DataFrame([
        {"User ID": 1,  "Level": "FC1", "Construction": 1.0,  "Research": 1.0,  "Troops": 1.0,  "FCs": 1200, "FC Shards": 210, "Time UTC": "16:00,16:30,17:00,17:30,18:00,18:30", "Days": "1,2"},
        {"User ID": 2,  "Level": "FC2", "Construction": 2.08, "Research": 2.0,  "Troops": 2.08, "FCs": 2100, "FC Shards": 380, "Time UTC": "07:00,07:30,08:00,08:30,09:00,09:30,10:00,10:30,11:00,11:30,12:00,12:30,13:00,13:30,14:00,14:30,15:00,15:30,16:00,16:30,17:00,17:30,18:00,18:30,19:00,19:30,20:00,20:30", "Days": "2,4"},
        {"User ID": 3,  "Level": "FC3", "Construction": 3.0,  "Research": 3.04, "Troops": 3.0,  "FCs": 2450, "FC Shards": 290, "Time UTC": "00:00,00:30,10:00,10:30,11:00,11:30,23:00,23:30", "Days": "1,2"},
        {"User ID": 4,  "Level": "FC4", "Construction": 4.0,  "Research": 4.17, "Troops": 4.0,  "FCs": 2600, "FC Shards": 410, "Time UTC": "14:00,14:30,15:00,15:30,16:00,16:30", "Days": "1,2,4"},
        {"User ID": 5,  "Level": "FC5", "Construction": 5.0,  "Research": 5.21, "Troops": 5.0,  "FCs": 2693, "FC Shards": 434, "Time UTC": "00:00,00:30,01:00,01:30,02:00,02:30,03:00,03:30,09:00,09:30,20:00,20:30,21:00,21:30,22:00,22:30,23:00,23:30", "Days": "1,4"},
        {"User ID": 6,  "Level": "FC2", "Construction": 2.5,  "Research": 2.3,  "Troops": 2.1,  "FCs": 1950, "FC Shards": 310, "Time UTC": "16:00,16:30,17:00,17:30", "Days": "1"},
        {"User ID": 7,  "Level": "FC3", "Construction": 3.3,  "Research": 3.1,  "Troops": 2.8,  "FCs": 2300, "FC Shards": 350, "Time UTC": "16:00,16:30,17:00,17:30", "Days": "1"},
        {"User ID": 8,  "Level": "FC1", "Construction": 1.1,  "Research": 0.9,  "Troops": 1.0,  "FCs": 900,  "FC Shards": 150, "Time UTC": "16:00,16:30,17:00,17:30", "Days": "1"},
        {"User ID": 9,  "Level": "FC2", "Construction": 2.0,  "Research": 1.8,  "Troops": 1.9,  "FCs": 1800, "FC Shards": 270, "Time UTC": "16:00,16:30,17:00,17:30", "Days": "1"},
        {"User ID": 10, "Level": "FC4", "Construction": 4.5,  "Research": 4.2,  "Troops": 4.3,  "FCs": 2550, "FC Shards": 420, "Time UTC": "16:00,16:30,17:00,17:30", "Days": "1"},
        {"User ID": 11, "Level": "FC3", "Construction": 3.7,  "Research": 3.5,  "Troops": 3.4,  "FCs": 2400, "FC Shards": 390, "Time UTC": "16:00,16:30,17:00,17:30", "Days": "1,2"},
        {"User ID": 12, "Level": "FC2", "Construction": 2.2,  "Research": 2.4,  "Troops": 2.0,  "FCs": 2050, "FC Shards": 330, "Time UTC": "06:00,06:30,07:00,07:30,08:00,08:30", "Days": "2,4"},
        {"User ID": 13, "Level": "FC5", "Construction": 5.3,  "Research": 5.1,  "Troops": 5.2,  "FCs": 2700, "FC Shards": 440, "Time UTC": "12:00,12:30,13:00,13:30,14:00,14:30", "Days": "1,2,4"},
        {"User ID": 14, "Level": "FC1", "Construction": 1.3,  "Research": 1.1,  "Troops": 1.4,  "FCs": 1100, "FC Shards": 190, "Time UTC": "22:00,22:30,23:00,23:30", "Days": "2"},
        {"User ID": 15, "Level": "FC4", "Construction": 4.1,  "Research": 3.9,  "Troops": 4.0,  "FCs": 2580, "FC Shards": 415, "Time UTC": "16:00,16:30,17:00,17:30", "Days": "1"},
    ])

    st.markdown('<div class="section-label">Data source</div>', unsafe_allow_html=True)

    source_options = ["Use parsed data from Parser", "Use built-in sample data", "Upload CSV / Excel"]
    default_src    = 0 if st.session_state["parsed_df"] is not None else 1
    source = st.radio(
        "Data source", source_options, index=default_src,
        horizontal=True, label_visibility="collapsed",
    )

    raw_df = None
    if source == "Use parsed data from Parser":
        if st.session_state["parsed_df"] is not None:
            raw_df = st.session_state["parsed_df"].copy()
            for col in ["Construction", "Research", "Troops"]:
                raw_df[col] = pd.to_numeric(raw_df[col], errors="coerce")
            n_before = len(raw_df)
            raw_df   = raw_df.dropna(subset=["Construction", "Research", "Troops"])
            n_dropped = n_before - len(raw_df)
            banner(
                "success",
                f"✓ Using {len(raw_df)} player(s) from Parser"
                + (f" — {n_dropped} skipped (missing numeric fields)" if n_dropped else ""),
            )
        else:
            banner("alert",
                   'No parsed data yet — go to the Parser page first, then click "Send to Scheduler".')

    elif source == "Use built-in sample data":
        raw_df = SCHEDULER_SAMPLE.copy()
        banner("info", "ℹ Using built-in sample dataset (15 users).")

    else:
        upload = st.file_uploader("Upload file", type=["csv", "xlsx"])
        if upload:
            raw_df = (
                pd.read_csv(upload) if upload.name.endswith(".csv")
                else pd.read_excel(upload)
            )
            banner("success", f"✓ Loaded {len(raw_df)} rows.")

    if raw_df is not None:
        with st.expander("Preview loaded data", expanded=False):
            st.dataframe(raw_df, use_container_width=True)

    # ── Manual slot overrides — always visible ────────────────────────────────
    st.markdown("<hr>", unsafe_allow_html=True)
    _override_player_ids = (
        [str(row[raw_df.columns[0]]) for _, row in raw_df.iterrows()]
        if raw_df is not None else []
    )
    with st.expander("📌 Manual slot overrides", expanded=bool(st.session_state["slot_overrides"])):
        if not _override_player_ids:
            banner("info", "ℹ Load data using the <b>Data source</b> selector above to enable player pins.")
        else:
            st.caption(
                "Pin one or more players to a specific slot before the solver runs. "
                "The solver will treat these pins as pre-filled and work around them. "
                "Leave a day blank to let the solver decide freely."
            )
            override_uid = st.selectbox(
                "Select player to pin",
                options=["— none —"] + _override_player_ids,
                key="override_uid_select",
            )
            if override_uid != "— none —":
                st.markdown(f"**Pinning User {override_uid}**")
                ov_cols = st.columns(len(DAY_CONFIG))
                for ci, dc in enumerate(DAY_CONFIG):
                    with ov_cols[ci]:
                        all_slot_labels = ["(solver decides)"] + [slot_label(s) for s in range(48)]
                        current_pin = st.session_state["slot_overrides"].get(override_uid, {}).get(dc["day"])
                        current_idx = (current_pin + 1) if current_pin is not None else 0
                        sel = st.selectbox(
                            dc["label"],
                            options=all_slot_labels,
                            index=current_idx,
                            key=f"ov_{override_uid}_{dc['day']}",
                        )
                        uid_ovs = st.session_state["slot_overrides"].setdefault(override_uid, {})
                        if sel == "(solver decides)":
                            uid_ovs.pop(dc["day"], None)
                        else:
                            uid_ovs[dc["day"]] = all_slot_labels.index(sel) - 1

        # Active pins table — shown even when no data loaded (so saved pins are visible)
        active_overrides = {
            uid: days for uid, days in st.session_state["slot_overrides"].items() if days
        }
        if active_overrides:
            st.markdown("**Active pins:**")
            pin_rows = []
            for uid, days in active_overrides.items():
                for day, slot in days.items():
                    dc_label = next((d["label"] for d in DAY_CONFIG if d["day"] == day), f"Day {day}")
                    pin_rows.append({"User ID": uid, "Day": dc_label, "Pinned slot": slot_label(slot)})
            st.dataframe(pd.DataFrame(pin_rows), use_container_width=True, hide_index=True)
            if st.button("🗑 Clear all pins", key="clear_pins"):
                st.session_state["slot_overrides"] = {}
                st.rerun()

    if raw_df is not None:
        st.markdown(
            '<div class="section-label" style="margin-top:1rem">Column mapping</div>',
            unsafe_allow_html=True,
        )
        cols = raw_df.columns.tolist()

        def pick(label: str, default: str) -> str:
            return st.selectbox(label, cols, index=cols.index(default) if default in cols else 0)

        c1, c2, c3 = st.columns(3)
        with c1:
            id_col  = pick("User ID column",     "User ID")
            lvl_col = pick("Level column",        "Level")
            fcs_col = pick("FCs column",          "FCs")
        with c2:
            con_col    = pick("Construction column", "Construction")
            res_col    = pick("Research column",     "Research")
            trp_col    = pick("Troops column",       "Troops")
            shards_col = pick("FC Shards column",    "FC Shards")
        with c3:
            time_default = next(
                (c for c in ("Time", "Hours", "Time UTC") if c in cols), cols[0]
            )
            time_col = pick("Time UTC / Hours column", time_default)
            days_col = pick("Days column", "Days")

        st.markdown("<hr>", unsafe_allow_html=True)

        banner("info", """
            <b>Scheduling objective:</b> Maximise the number of players assigned a slot.
            Among all solutions that place the same number of players, the highest-scoring
            players are preferred — guaranteed by min-cost max-flow (provably optimal score ranking).
            Any remaining unplaced high-scorers are mathematically unplaceable: their entire
            time window is saturated by even-higher-scorers with no alternative slots.
        """)

        run_label_input = st.text_input(
            "Run label (optional)",
            placeholder="e.g. 'After fixing User 42 window'",
            key="run_label_input",
        )

        if st.button("⚡ Run scheduler", type="primary", use_container_width=False):
            def _safe_int(row, col_name):
                """Safely coerce a DataFrame cell to int, returning None on failure."""
                val = row[col_name]
                try:
                    return int(float(val)) if pd.notna(val) and str(val).strip() not in ("", "—") else None
                except (ValueError, TypeError):
                    return None

            users = []
            for _, row in raw_df.iterrows():
                raw_time_val = str(row[time_col]) if pd.notna(row[time_col]) else ""
                hours = slots_str_to_hours(raw_time_val) if ":" in raw_time_val else parse_ints(raw_time_val)

                try:
                    con_val = float(row[con_col])
                    res_val = float(row[res_col])
                    trp_val = float(row[trp_col])
                except (ValueError, TypeError):
                    continue

                users.append({
                    "User ID":      row[id_col],
                    "Level":        str(row[lvl_col]),
                    "Construction": con_val,
                    "Research":     res_val,
                    "Troops":       trp_val,
                    "FCs":          _safe_int(row, fcs_col),
                    "FC Shards":    _safe_int(row, shards_col),
                    "hours":        hours,
                    "days":         parse_ints(row[days_col]),
                })

            if not users:
                st.error("No valid users to schedule — check your column mapping and data.")
            else:
                # Apply manual overrides: inject pins into slot_occ before the solver
                # by adding a synthetic "already placed" signal via hours restriction.
                # Strategy: for pinned users on a given day, replace their hours with
                # only the pinned hour so MCMF is forced to that slot.
                overrides = st.session_state["slot_overrides"]
                users_for_run = []
                for u in users:
                    uid_str = str(u["User ID"])
                    if uid_str not in overrides:
                        users_for_run.append(u)
                        continue
                    u_copy = dict(u)
                    # We'll apply per-day overrides inside run_day via a patched copy.
                    # Store the override map on the user dict for run_day to read.
                    u_copy["_slot_overrides"] = overrides[uid_str]
                    users_for_run.append(u_copy)

                with st.spinner("Running min-cost max-flow scheduler…"):
                    day_results = run_scheduler(users_for_run)
                    # Day 4 — VP: second pass for players unassigned in Day 4 MoE
                    day4_result = next((dr for dr in day_results if dr["day"] == 4), None)
                    vp4_result  = run_day4_vp(users_for_run, day4_result) if day4_result else None

                # ── Save to run history ────────────────────────────────────────
                run_ts    = datetime.utcnow().strftime("%H:%M UTC")
                run_label = run_label_input.strip() or run_ts
                # Snapshot: {uid: {day: slot_index}} for every result
                snap: dict = {}
                for dr in day_results:
                    for uid, slot in dr["user_slot"].items():
                        snap.setdefault(str(uid), {})[dr["day"]] = slot
                if vp4_result:
                    for uid, slot in vp4_result["user_slot"].items():
                        snap.setdefault(str(uid), {})["4vp"] = slot

                history = st.session_state["run_history"]
                history.append({"label": run_label, "snap": snap, "ts": run_ts})
                # Keep last 5 runs
                st.session_state["run_history"] = history[-5:]

                st.markdown('<div class="section-label" style="margin-top:1.5rem">Results</div>',
                            unsafe_allow_html=True)

                total_possible   = sum(len(dr["eligible"])  for dr in day_results)
                total_assigned   = sum(len(dr["user_slot"]) for dr in day_results)
                total_unassigned = sum(len(dr["unassigned"]) for dr in day_results)
                pct = round(100 * total_assigned / total_possible) if total_possible else 0

                stat_row(
                    stat_card(total_possible,   "Eligible slots-days"),
                    stat_card(total_assigned,   "Assigned"),
                    stat_card(f"{pct}%",        "Fill rate"),
                    stat_card(total_unassigned, "Unassignable", warn=bool(total_unassigned)),
                )

                # ── Changelog / diff view ──────────────────────────────────────
                if len(st.session_state["run_history"]) >= 2:
                    history = st.session_state["run_history"]
                    with st.expander("🔄 Changelog — diff vs previous run", expanded=False):
                        compare_options = [h["label"] for h in history[:-1]]
                        compare_label   = st.selectbox(
                            "Compare current run against:",
                            options=compare_options,
                            index=len(compare_options) - 1,
                            key="changelog_compare_select",
                        )
                        prev_snap = next(h["snap"] for h in history if h["label"] == compare_label)
                        curr_snap = history[-1]["snap"]

                        day_keys = {dc["day"]: dc["label"] for dc in DAY_CONFIG}
                        day_keys["4vp"] = DAY4_VP_CONFIG["label"]

                        all_uids = sorted(set(prev_snap) | set(curr_snap), key=str)
                        diff_rows = []
                        for uid in all_uids:
                            prev_days = prev_snap.get(uid, {})
                            curr_days = curr_snap.get(uid, {})
                            all_days  = set(prev_days) | set(curr_days)
                            for day in sorted(all_days, key=lambda d: (str(d))):
                                p = prev_days.get(day)
                                c = curr_days.get(day)
                                if p == c:
                                    continue
                                diff_rows.append({
                                    "User ID":  uid,
                                    "Day":      day_keys.get(day, f"Day {day}"),
                                    "Previous": slot_label(p) if p is not None else "unassigned",
                                    "Current":  slot_label(c) if c is not None else "unassigned",
                                    "Change":   (
                                        "✅ now assigned"   if p is None and c is not None else
                                        "❌ now unassigned" if p is not None and c is None else
                                        "↔ moved slot"
                                    ),
                                })

                        if not diff_rows:
                            banner("success", "✓ No changes — this run produced identical assignments.")
                        else:
                            n_better = sum(1 for r in diff_rows if r["Change"] == "✅ now assigned")
                            n_worse  = sum(1 for r in diff_rows if r["Change"] == "❌ now unassigned")
                            n_moved  = sum(1 for r in diff_rows if r["Change"] == "↔ moved slot")
                            banner(
                                "info",
                                f"<b>{len(diff_rows)} change(s)</b> vs <em>{compare_label}</em>: "
                                f"{n_better} newly assigned &nbsp;·&nbsp; "
                                f"{n_worse} newly unassigned &nbsp;·&nbsp; "
                                f"{n_moved} slot change(s).",
                            )
                            st.dataframe(
                                pd.DataFrame(diff_rows),
                                use_container_width=True,
                                hide_index=True,
                                column_config={
                                    "User ID":  st.column_config.TextColumn(width="small"),
                                    "Day":      st.column_config.TextColumn(width="medium"),
                                    "Previous": st.column_config.TextColumn(width="medium"),
                                    "Current":  st.column_config.TextColumn(width="medium"),
                                    "Change":   st.column_config.TextColumn(width="medium"),
                                },
                            )

                 # --- 📊 GLOBAL AVAILABILITY CHART ---
                st.markdown("#### 📊 Peak Availability")
                st.caption("Total number of players available at each 30-minute time slot.")
                
                # Count available players per slot
                slot_counts = {s: 0 for s in range(48)}
                for u in users_for_run:
                    for s in _user_slots(u):
                        slot_counts[s] += 1
                        
                # Format for Streamlit chart
                chart_df = pd.DataFrame([
                    {"Time Slot": slot_label(s), "Players Available": count}
                    for s, count in slot_counts.items()
                ])
                
                # Display a cyan bar chart matching your theme
                st.bar_chart(chart_df.set_index("Time Slot"), color="#00e5cc")

                st.markdown("#### 📋 User summary — all days")
                st.caption(
                    "Each cell shows the assigned slot and speedups. "
                    "❌ = window saturated (all their slots taken by higher-speedup players). "
                    "— = not participating on that day."
                )
                st.dataframe(build_summary_df(users_for_run, day_results, vp4_result), use_container_width=True, hide_index=True)

                st.markdown("#### 📅 Per-day detail")
                tab_labels = [dc["label"] for dc in DAY_CONFIG]
                if vp4_result is not None:
                    tab_labels.append(DAY4_VP_CONFIG["label"])
                tabs = st.tabs(tab_labels)
                for i, dc in enumerate(DAY_CONFIG):
                    dr = day_results[i]
                    with tabs[i]:
                        ca, cb, cc = st.columns(3)
                        ca.metric("Eligible users", len(dr["eligible"]))
                        cb.metric("Assigned",       len(dr["user_slot"]))
                        cc.metric("Unassigned",     len(dr["unassigned"]))

                        if dr["user_slot"] and dr["unassigned"]:
                            placed_scores   = [u[dc["col"]] for u in dr["eligible"] if u["User ID"] in dr["user_slot"]]
                            unplaced_scores = [e["user"][dc["col"]] for e in dr["unassigned"]]
                            min_p = min(placed_scores)
                            max_u = max(unplaced_scores)
                            if max_u > min_p + 0.001:
                                banner(
                                    "info",
                                    f"ℹ Some unassigned players have more speedups than the "
                                    f"lowest-placed player (highest unassigned: <b>{max_u:.2f}</b>, "
                                    f"lowest placed: <b>{min_p:.2f}</b>).",
                                )
                            else:
                                banner(
                                    "success",
                                    f"✓ Optimal speedup order: every placed player has ≥ speedups "
                                    f"than every unplaced player "
                                    f"(min placed {min_p:.2f} ≥ max unplaced {max_u:.2f}).",
                                )

                        st.markdown("**Slot timeline**")
                        extra_col = _DAY_EXTRA_COL.get(dc["day"])
                        timeline_df = build_timeline_df(users_for_run, dr, extra_col=extra_col)
                        col_cfg = {
                            "Speedups":  st.column_config.TextColumn(width="small"),
                            "Time Slot": st.column_config.TextColumn(width="medium"),
                            "Assigned":  st.column_config.TextColumn(width="medium"),
                        }
                        if extra_col:
                            col_cfg[extra_col] = st.column_config.TextColumn(extra_col, width="small")
                        st.dataframe(timeline_df, use_container_width=True, column_config=col_cfg)

                        ua_df = build_unassigned_df(dr)
                        if ua_df.empty:
                            banner("success", "✅ All eligible users assigned on this day.")
                        else:
                            st.markdown("**Unassigned users**")
                            st.dataframe(ua_df, use_container_width=True, hide_index=True)

                # ── Day 4 — VP tab ────────────────────────────────────────────────────
                if vp4_result is not None:
                    with tabs[len(DAY_CONFIG)]:
                        banner(
                            "info",
                            "ℹ <b>Day 4 — VP</b> gives players who were <b>unassigned in Day 4 — MoE</b> "
                            "a second-chance VP slot.",
                        )

                        vp4_eligible   = vp4_result["eligible"]
                        vp4_assigned   = vp4_result["user_slot"]
                        vp4_unassigned = vp4_result["unassigned"]

                        ca, cb, cc = st.columns(3)
                        ca.metric("Eligible (unassigned from MoE)", len(vp4_eligible))
                        cb.metric("Assigned VP slots",              len(vp4_assigned))
                        cc.metric("Still unassigned",               len(vp4_unassigned))

                        if not vp4_eligible:
                            banner("success", "✅ No unassigned players from Day 4 — MoE. Nothing to schedule.")
                        else:
                            if vp4_assigned and vp4_unassigned:
                                placed_scores   = [u["Troops"] for u in vp4_eligible if u["User ID"] in vp4_assigned]
                                unplaced_scores = [e["user"]["Troops"] for e in vp4_unassigned]
                                min_p = min(placed_scores)
                                max_u = max(unplaced_scores)
                                if max_u > min_p + 0.001:
                                    banner(
                                        "info",
                                        f"ℹ Some unassigned players have more speedups than the "
                                        f"lowest-placed player (highest unassigned: <b>{max_u:.2f}</b>, "
                                        f"lowest placed: <b>{min_p:.2f}</b>). "
                                        f"Their entire time window is saturated — the scheduler cannot "
                                        f"free a slot without dropping someone else.",
                                    )
                                else:
                                    banner(
                                        "success",
                                        f"✓ Optimal speedup order: every placed player has ≥ speedups "
                                        f"than every unplaced player "
                                        f"(min placed {min_p:.2f} ≥ max unplaced {max_u:.2f}).",
                                    )

                            st.markdown("**Slot timeline**")
                            vp4_timeline_df = build_timeline_df(users_for_run, vp4_result)
                            st.dataframe(
                                vp4_timeline_df,
                                use_container_width=True,
                                column_config={
                                    "Speedups":  st.column_config.TextColumn(width="small"),
                                    "Time Slot": st.column_config.TextColumn(width="medium"),
                                    "Assigned":  st.column_config.TextColumn(width="medium"),
                                },
                            )

                            vp4_ua_df = build_unassigned_df(vp4_result)
                            if vp4_ua_df.empty:
                                banner("success", "✅ All eligible players assigned a VP slot.")
                            else:
                                st.markdown("**Still unassigned after VP pass**")
                                st.dataframe(vp4_ua_df, use_container_width=True, hide_index=True)

                st.markdown("<hr>", unsafe_allow_html=True)
                st.markdown('<div class="section-label">Export</div>', unsafe_allow_html=True)
                st.download_button(
                    label="📥 Download full schedule (.xlsx)",
                    data=to_excel_schedule(users_for_run, day_results, vp4_result),
                    file_name="SvS_schedule.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
